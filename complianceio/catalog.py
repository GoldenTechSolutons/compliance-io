import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from uuid import UUID, uuid4

import click
import openpyxl
from trestle.oscal.catalog import Catalog, Control, Group, Model
from trestle.oscal.common import Link, Metadata, Part, Property

from oscal.oscal import control_to_statement_id


def create_groups(p):
    """
    Parse the spreadsheet to get groups by family.
    """
    with open(Path.cwd().joinpath('complianceio', 'header_map.json'), 'r') as f:
        c_map = json.load(f)
    wb = openpyxl.load_workbook(filename=p, data_only=True)
    groups = []
    ws = wb.worksheets[1]
    header = [cell for cell in ws['A2:XFD2'][0] if cell.value is not None and cell.value.strip() != '']
    family_id = None
    group = None
    controls = []
    bline = {}
    for row in ws.iter_rows(min_row=3, max_col=19, max_row=1100):
        if row[0].value is not None:
            family = row[c_map.get('family')].value
            control_id = row[c_map.get('control_id')].value
            group = control_id[:2].lower()
            name = row[c_map.get('name')].value
            control_text = row[c_map.get('control')].value
            implementation = row[c_map.get('implementation')].value
            responsiblity = row[c_map.get('responsiblity')].value
            related = row[c_map.get('responsiblity')].value
            reference = row[c_map.get('reference')].value
            baseline = row[c_map.get('baseline')].value

            if family_id is not None and family != family_id:
                g = add_group(group_id, family_id, controls)
                groups.append(g)
                controls = []

            if control_text is not None:
                cid = control_to_statement_id(control_id)
                parts = []
                part = parse_parts(control_text, cid)
                parts.append(part)

                imp = row[c_map.get('implementation')].value
                if imp and imp.strip():
                    pid = cid.replace('_smt', '_imp')
                    part = add_additional(pid, 'implementation', imp.strip())
                    parts.append(part)

                hva = row[c_map.get('hva_standards')].value
                if hva and hva.strip():
                    pid = cid.replace('_smt', '_hva')
                    part = add_additional(pid, 'hva', hva.strip())
                    parts.append(part)

                prv = row[c_map.get('privacy_standards')].value
                if prv and prv.strip():
                    pid = cid.replace('_smt', '_prv')
                    part = add_additional(pid, 'privacy', prv.strip())
                    parts.append(part)

                gdn = row[c_map.get('discussion')].value
                if gdn and gdn.strip():
                    pid = cid.replace('_smt', '_gdn')
                    part = add_additional(pid, 'guidance', gdn.strip())
                    parts.append(part)

                links = []
                related = row[c_map.get('related')].value
                if related and related.strip():
                    rlt = related.strip().replace(' ', ',').split(',')
                    for r in rlt:
                        if r[:4] != 'None':
                            link = r.strip()
                            if link:
                                links.append(Link(
                                    href=f'#{link}',
                                    rel='related'
                                ))
                l = links if len(links) > 0 else None

                controls.append(Control(
                    id=cid.strip('_smt'),
                    class_='ARS-5.0-Mandatory',
                    title=name,
                    props=[Property(
                        name='label',
                        value=control_id
                    ), Property(
                        name='sort-id',
                        value=control_id.lower()
                    )],
                    links=l,
                    parts=parts,
                ))
        family_id = family
        group_id = group

    g = add_group(group_id, family, controls)
    groups.append(g)
    return groups


def add_group(group_id, family, controls):
    return Group(
        id=group_id,
        class_='family',
        title=family,
        controls=controls
    )


def parse_parts(text, cid):
    """
    This is very specific to the CMS spreadsheet.
    """
    pt = parse_control(text, cid)
    parts = []
    if pt:
        parts = create_parts(pt)

    part = Part(
        id=cid,
        name='statement',
        parts=parts,
    )

    return part


def parse_control(text, cid):
    """
    This is looking for a very specifically formatted control statement.
    Top level parts should start with (LETTER), for instance (a)
    Secondary level should start with NUMBER. for instance 1.
    Tertiary level should start with LETTER. for instance a.
    As an example:
    (a) This is the top level
        1. With some subtext
            a. And tertiary text.
    """
    lines = text.splitlines()
    parts = {cid.strip(): {'prose': None, 'children': {}}}
    pr = None
    for l in lines:
        text = l.strip()
        if re.search(r'^\(([a-z])\)', text):
            first = re.search(r'^\(([a-z])\)', text)
            if first:
                f = first.group()
                fid = f.replace('(', '').replace(')', '')
                prose = text.strip(f'({f}) ')
                parts[cid.strip()]['children'][fid.strip()] = {
                    'prose': prose,
                    'children': {},
                }
        elif re.search(r'^[1-9]\.\s', text):
            second = re.search(r'^[1-9]\.\s', text)
            if second:
                s = second.group()
                sid = s.replace('. ', '')
                prose = text.strip(f'{s}')
                parts[cid.strip()]['children'][fid.strip()]['children']\
                    [sid.strip()] = {
                    'prose': prose,
                    'children': {},
                }
        elif re.search(r'^[a-z]\.\s', text):
            third = re.search(r'^[a-z]\.\s', text)
            if third:
                t = third.group()
                tid = t.replace('. ', '')
                prose = text.strip(f'{t}')
                parts[cid.strip()]['children'][fid.strip()]\
                    ['children'][sid.strip()]['children'][tid.strip()] = {'prose': prose}
        else:
            pr = f'{pr}\n{l}' if pr else l
            parts[cid.strip()]['prose'] = pr

    return parts


def create_parts(parts):
    part = []
    first = None
    secondary = None
    tertiary = None
    for k, pv in parts.items():
        if 'children' in pv:
            first = []
            for f, fv in pv.get('children').items():
                if 'children' in fv:
                    secondary = []
                    for s, sv in fv.get('children').items():
                        if 'children' in sv:
                            tertiary = []
                            for t, tv in sv.get('children').items():
                                tertiary.append(Part(
                                    id=f'{k}.{f}.{s}.{t}',
                                    name='item',
                                    props=[Property(
                                        name='label',
                                        value=t,
                                    )],
                                    prose=tv.get('prose')
                                ))
                        tp = tertiary if len(tertiary) > 0 else None
                        secondary.append(Part(
                            id=f'{k}.{f}.{s}',
                            name='item',
                            props=[Property(
                                name='label',
                                value=s,
                            )],
                            parts=tp,
                            prose=sv.get('prose'),
                        ))
                sp = secondary if len(secondary) > 0 else None
                first.append(Part(
                    id=f'{k}.{f}',
                    name='item',
                    props=[Property(
                        name='label',
                        value=f,
                    )],
                    parts=sp,
                    prose=fv.get('prose'),
                ))
        fp = first if len(first) > 0 else None
        part.append(Part(
            id=f'{k}',
            name='item',
            props=[Property(
                name='label',
                value=k,
            )],
            parts=fp,
            prose=pv.get('prose'),
        ))
    return part


def add_additional(pid, name, prose):
    return Part(
        id=pid,
        name=name,
        prose=prose,
    )


@click.command()
@click.option('-t', '--title', required=True)
@click.argument(
    'source',
    type=click.Path(exists=True, dir_okay=False, file_okay=True, resolve_path=True)

)
def main(title, source):
    """
    Create an OSCAL Catalog.
    """
    p = Path(source)
    groups = create_groups(p)
    uuid = uuid4().urn

    today = datetime.now(timezone(timedelta(hours=-6))).isoformat()

    md = Metadata(
        title=title,
        last_modified=str(today),
        version='1.0',
        oscal_version='1.0.0'
    )

    catalog = Catalog(
            uuid=uuid[9:],
            metadata=md,
            groups=groups,
        )
    root = Model(catalog=catalog).dict(by_alias=True, exclude_unset=True, exclude_none=True)
    print(json.dumps(root, indent=2, default=str))

if __name__ == "__main__":
    main()
